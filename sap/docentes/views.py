from django.forms import modelform_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.template import loader
from openpyxl.workbook import Workbook

from docentes.forms import DocenteFormulario
from docentes.models import Docente


# Create your views here.

def agregar_docente(request):
    pagina = loader.get_template('agregar_docentes.html')
    if request.method == 'GET':
         formulario = DocenteFormulario
    elif request.method == 'POST':
        formulario = DocenteFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_docente(request, idDocente):
    pagina = loader.get_template('ver_docentes.html')
    docente = get_object_or_404(Docente, pk=idDocente)
    mensaje = {'docente': docente}
    return HttpResponse(pagina.render(mensaje, request))



def editar_docente(request, idDocente):
    pagina = loader.get_template('editar_docentes.html')
    docente = get_object_or_404(Docente, pk=idDocente)
    if request.method == "GET":
        formulario = DocenteFormulario(instance=docente)

    elif request.method == "POST":
        formulario = DocenteFormulario(request.POST, instance=docente)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')

    mensaje = {'formulario': formulario}

    return HttpResponse(pagina.render(mensaje, request))

def eliminar_docente(request, idDocente):
    docente = get_object_or_404(Docente, pk=idDocente)
    if docente:
        docente.delete()
        return redirect('inicio')
from django.shortcuts import render

# Create your views here.
def generar_reporte(request):
    # docentes = Docente.objects.all()
    docentes = Docente.objects.order_by('apellido')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE DOCENTES'
    ws['B1'] = 'REPORTE DE DOCENTES'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:H1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'APELLIDO'
    ws['D3'] = 'SEXO'
    ws['E3'] = 'EMAIL'
    ws['F3'] = 'MATERIA'
    ws['G3'] = 'UNIVERSIDAD'
    ws['H3'] = 'JORNADA'

    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for docente in docentes:
        ws.cell(row=cont, column=2).value = docente.nombre
        ws.cell(row=cont, column=3).value = docente.apellido
        ws.cell(row=cont, column=4).value = docente.sexo
        ws.cell(row=cont, column=5).value = docente.email
        ws.cell(row=cont, column=6).value = docente.materia.nombre
        ws.cell(row=cont, column=7).value = docente.universidad.nombre
        ws.cell(row=cont, column=8).value = docente.jornada
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReporteDocentesExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response